# surveybot/writer.py
# Takes flattened survey results and writes them to the Excel template

import openpyxl
from openpyxl.styles import PatternFill, Font
from copy import copy
from pathlib import Path
# from config import SURVEY_SOURCE
SURVEY_SOURCE = "S"

# Yellow fill for flagged cells (low confidence)
FLAG_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Excel column order (must match template exactly)
# Columns with empty header in template (Q9, Q10, Q11 group headers) are skipped
DATA_COLUMNS = [
    ("A", "#"),
    ("B", "L_U"),
    ("C", "Q1"),
    ("D", "Q2"),
    ("E", "Q3"),
    ("F", "Q4"),
    ("G", "Q5"),
    ("H", "Q6"),
    ("I", "Q7"),
    ("J", "Q8"),
    ("K", None),      # Q9 group header - blank
    ("L", "Q9a"),
    ("M", "Q9b"),
    ("N", "Q9c"),
    ("O", "Q9d"),
    ("P", "Q9e"),
    ("Q", "Q9f"),
    ("R", None),      # Q10 group header - blank
    ("S", "Q10a"),
    ("T", "Q10b"),
    ("U", "Q10c"),
    ("V", "Q10d"),
    ("W", "Q10e"),
    ("X", "Q10f"),
    ("Y", "Q10g"),
    ("Z", None),      # Q11 group header - blank
    ("AA", "Q11a"),
    ("AB", "Q11b"),
    ("AC", "Q11c"),
    ("AD", "Q11d"),
    ("AE", "Q12"),
    ("AF", "Q13"),
    ("AG", "Q14"),
    ("AH", "Q15"),
    ("AI", "Q16"),
    ("AJ", "Q17"),
    ("AK", "source"),
]


def find_next_empty_row(ws, header_rows: int = 3) -> int:
    """
    Find the first empty data row after the header rows.
    Template has 3 header rows. Data starts at row 4.
    We append after the last populated row.
    """
    for row in range(header_rows + 1, ws.max_row + 10):
        if ws.cell(row=row, column=2).value is None:
            return row
    return ws.max_row + 1


def write_results(
    results: list[dict],
    template_path: str,
    output_path: str,
    start_number: int = None
) -> str:
    """
    Write extracted survey results into a copy of the Excel template.
    
    Args:
        results: List of flattened survey dicts from extractor.flatten_result()
        template_path: Path to the master Excel template
        output_path: Where to save the output file
        start_number: Override the survey number (if None, reads from survey_id)
    
    Returns:
        Path to the output file
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Find where to start writing (first empty data row)
    start_row = find_next_empty_row(ws)
    print(f"Writing to sheet '{ws.title}', starting at row {start_row}")

    for i, flat in enumerate(results):
        row = start_row + i
        is_error = flat.get("_status") == "error"
        flagged_keys = set(flat.get("_flags", "").split(",")) if flat.get("_flags") else set()

        # Determine survey number
        # if start_number is not None:
        #     # survey_num = start_number + i
        #     try:
        #         survey_num = int(sid.split("-")[1]) if sid else row - start_row + 1
        #     except (IndexError, ValueError):
        #         survey_num = row - start_row + 1
        # else: 
        #     # Parse from survey_id e.g. "L-1" → 1
        #     sid = flat.get("survey_id", "")
        #     try:
        #         survey_num = int(sid.split("-")[1])
        #     except (IndexError, ValueError):
        #         survey_num = row - start_row + 1
        # update cs 6/3/26
        if start_number is not None:
            survey_num = start_number + i  # just use this, simple
        else:
            sid = flat.get("survey_id", "")
            try:
                survey_num = int(sid.split("-")[1]) if sid else row - start_row + 1
            except (IndexError, ValueError):
                survey_num = row - start_row + 1

        for col_letter, field_key in DATA_COLUMNS:
            cell = ws[f"{col_letter}{row}"]

            if field_key is None:
                # Group header column — leave blank
                continue
            elif field_key == "#":
                cell.value = survey_num
            elif field_key == "L_U":
                cell.value = flat.get("likely_voter", "")
            elif field_key == "source":
                cell.value = SURVEY_SOURCE
            else:
                cell.value = flat.get(field_key)

            # Highlight flagged or error cells
            if is_error:
                cell.fill = copy(ERROR_FILL)
            elif field_key in flagged_keys:
                cell.fill = copy(FLAG_FILL)

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Wrote {len(results)} surveys ({sum(1 for r in results if r.get('_status') == 'ok')} OK, "
          f"{sum(1 for r in results if r.get('_status') == 'error')} errors, "
          f"{sum(1 for r in results if r.get('_flags'))} flagged)")
    return output_path


if __name__ == "__main__":
    # Quick writer test with dummy data
    dummy = [
        {
            "survey_id": "L-1", "likely_voter": "L", "_status": "ok", "_flags": "",
            "Q1": 1, "Q2": 1, "Q3": 3, "Q4": 1, "Q5": 1, "Q6": "4,5", "Q7": 2,
            "Q8": 3,
            "Q9a": 5, "Q9b": 3, "Q9c": 4, "Q9d": 5, "Q9e": 3, "Q9f": 3,
            "Q10a": 5, "Q10b": 1, "Q10c": 5, "Q10d": 4, "Q10e": 3, "Q10f": 3, "Q10g": 5,
            "Q11a": 1, "Q11b": 2, "Q11c": 5, "Q11d": 3,
            "Q12": 2, "Q13": 3, "Q14": 2, "Q15": 1, "Q16": "1", "Q17": "2",
            "source": "S"
        }
    ]
    template = "/mnt/user-data/uploads/Deerfield_109_Mail_Survey_Data_Entry_Spreadsheet_Template_Final_11_24_2025_FINAL.xlsx"
    out = write_results(dummy, template, "/mnt/user-data/outputs/test_output.xlsx")
    print(f"Test complete: {out}")
