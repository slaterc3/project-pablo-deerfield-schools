#!/usr/bin/env python3
# combine.py — merge all batch result files for a client into one Excel
#
# Usage:
#   python combine.py --client east_maine
#   python combine.py --client deerfield
#
# Looks for files matching: files/*{client_pattern}*results*.xlsx
# Writes to: files/{Client}_Mail_Survey_Results_Combined.xlsx

import argparse
import glob
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
from pathlib import Path
from datetime import datetime

# Map client name to file search pattern and template path
CLIENT_CONFIG = {
    "east_maine": {
        "pattern":  "files/*East Maine*results*.xlsx",
        "template": "files/East Maine Mail Survey Data Entry Spreadsheet Template.xlsx",
        "output":   "files/East_Maine_Mail_Survey_Results_Combined.xlsx",
    },
    "deerfield": {
        "pattern":  "files/*Deerfield*results*.xlsx",
        "template": "files/Deerfield 109 Mail Survey Data Entry Spreadsheet Template Final 11.24.2025 FINAL.xlsx",
        "output":   "files/Deerfield_Mail_Survey_Results_Combined.xlsx",
    },
    "troy_30c": {
        "pattern":  "files/*Troy*results*.xlsx",
        "template": "files/Troy_30C_Mail_Survey_Data_Entry_Spreadsheet_v01.xlsx",
        "output":   "files/Troy_30C_Mail_Survey_Results_Combined.xlsx",
    },
    "weld_re1": {
        "pattern":  "files/*Weld*results*.xlsx",
        "template": "files/Weld 1 Mail Survey Data Entry Spreadsheet Template_2.xlsx",
        "output":   "files/Weld_RE1_Mail_Survey_Results_Combined.xlsx",
    },
}

FLAG_FILL  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")


def find_next_empty_row(ws, header_rows: int = 3) -> int:
    for row in range(header_rows + 1, ws.max_row + 10):
        if ws.cell(row=row, column=2).value is None:
            return row
    return ws.max_row + 1


def get_data_rows(filepath: str, header_rows: int = 3):
    """Extract all data rows from a results file."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = []
    fills = []
    for row_idx in range(header_rows + 1, ws.max_row + 1):
        row_data = []
        row_fills = []
        # Check if row has data
        if ws.cell(row=row_idx, column=2).value is None:
            break
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(cell.value)
            row_fills.append(copy(cell.fill) if cell.fill.fill_type != 'none' else None)
        rows.append(row_data)
        fills.append(row_fills)
    wb.close()
    return rows, fills


def main():
    parser = argparse.ArgumentParser(description="Combine batch result files into one Excel")
    parser.add_argument("--client", required=True, choices=list(CLIENT_CONFIG.keys()),
                        help="Client name")
    parser.add_argument("--header-rows", type=int, default=3,
                        help="Number of header rows in template (default: 3)")
    args = parser.parse_args()

    config = CLIENT_CONFIG[args.client]

    # Find all result files sorted by name
    result_files = sorted(glob.glob(config["pattern"]))
    if not result_files:
        print(f"No result files found matching: {config['pattern']}")
        return

    print(f"Found {len(result_files)} result files:")
    for f in result_files:
        print(f"  {f}")

    # Load template
    print(f"\nLoading template: {config['template']}")
    wb_out = openpyxl.load_workbook(config["template"])
    ws_out = wb_out.active

    # Find insertion point
    start_row = find_next_empty_row(ws_out, header_rows=args.header_rows)
    print(f"Writing combined results starting at row {start_row}")

    total_rows = 0
    for filepath in result_files:
        rows, fills = get_data_rows(filepath, header_rows=args.header_rows)
        print(f"  {Path(filepath).name}: {len(rows)} surveys")

        for i, (row_data, row_fills) in enumerate(zip(rows, fills)):
            out_row = start_row + total_rows + i
            for col_idx, (value, fill) in enumerate(zip(row_data, row_fills), start=1):
                cell = ws_out.cell(row=out_row, column=col_idx, value=value)
                if fill:
                    cell.fill = fill

        total_rows += len(rows)

    wb_out.save(config["output"])
    print(f"\nCombined {total_rows} surveys → {config['output']}")


if __name__ == "__main__":
    main()
