# surveybot/writer.py
# Takes flattened survey results and writes them to the Excel template
# Supports per-survey-type column mappings via surveys/<type>/config.py

import openpyxl
import importlib
from openpyxl.styles import PatternFill
from copy import copy
from pathlib import Path

SURVEY_SOURCE = "S"

FLAG_FILL  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Default Deerfield column mapping (fallback)
DEERFIELD_COLUMNS = [
    ("A", None),
    ("B", "#"),
    ("C", "L_U"),
    ("D", "Q1"),
    ("E", "Q2"),
    ("F", "Q3"),
    ("G", "Q4"),
    ("H", "Q5"),
    ("I", "Q6"),
    ("J", "Q7"),
    ("K", "Q8"),
    ("L", None),      # Q9 group header
    ("M", "Q9a"),
    ("N", "Q9b"),
    ("O", "Q9c"),
    ("P", "Q9d"),
    ("Q", "Q9e"),
    ("R", "Q9f"),
    ("S", None),      # Q10 group header
    ("T", "Q10a"),
    ("U", "Q10b"),
    ("V", "Q10c"),
    ("W", "Q10d"),
    ("X", "Q10e"),
    ("Y", "Q10f"),
    ("Z", "Q10g"),
    ("AA", None),     # Q11 group header
    ("AB", "Q11a"),
    ("AC", "Q11b"),
    ("AD", "Q11c"),
    ("AE", "Q11d"),
    ("AF", "Q12"),
    ("AG", "Q13"),
    ("AH", "Q14"),
    ("AI", "Q15"),
    ("AJ", "Q16"),
    ("AK", "Q17"),
    ("AL", "source"),
]


def load_header_rows(survey_type: str) -> int:
    """Load HEADER_ROWS from survey config, default to 3."""
    try:
        config = importlib.import_module(f"surveys.{survey_type}.config")
        return getattr(config, "HEADER_ROWS", 3)
    except Exception:
        return 3


def load_column_map(survey_type: str) -> list:
    """Load DATA_COLUMNS from survey config, fall back to Deerfield if not defined."""
    try:
        config = importlib.import_module(f"surveys.{survey_type}.config")
        return config.DATA_COLUMNS
    except AttributeError:
        return DEERFIELD_COLUMNS


def find_next_empty_row(ws, header_rows: int = 3) -> int:
    """Find first empty data row after header rows."""
    for row in range(header_rows + 1, ws.max_row + 10):
        if ws.cell(row=row, column=2).value is None:
            return row
    return ws.max_row + 1


def write_results(
    results: list[dict],
    template_path: str,
    output_path: str,
    start_number: int = None,
    survey_type: str = "deerfield"
) -> str:
    """Write extracted survey results into a copy of the Excel template."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    data_columns = load_column_map(survey_type)
    header_rows = load_header_rows(survey_type)

    start_row = find_next_empty_row(ws, header_rows=header_rows)
    print(f"Writing to sheet '{ws.title}', starting at row {start_row}")

    for i, flat in enumerate(results):
        row = start_row + i
        is_error = flat.get("_status") == "error"
        flagged_keys = set(flat.get("_flags", "").split(",")) if flat.get("_flags") else set()

        # Determine survey number
        if start_number is not None:
            survey_num = start_number + i
        else:
            sid = flat.get("survey_id", "")
            try:
                survey_num = int(sid.split("-")[1]) if sid else row - start_row + 1
            except (IndexError, ValueError):
                survey_num = row - start_row + 1

        for col_letter, field_key in data_columns:
            cell = ws[f"{col_letter}{row}"]

            if field_key is None:
                continue
            elif field_key == "#":
                cell.value = survey_num
            elif field_key == "L_U":
                cell.value = flat.get("likely_voter", "")
            elif field_key == "source":
                cell.value = SURVEY_SOURCE
            else:
                cell.value = flat.get(field_key)

            if is_error:
                cell.fill = copy(ERROR_FILL)
            elif field_key in flagged_keys:
                cell.fill = copy(FLAG_FILL)

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Wrote {len(results)} surveys "
          f"({sum(1 for r in results if r.get('_status') == 'ok')} OK, "
          f"{sum(1 for r in results if r.get('_status') == 'error')} errors, "
          f"{sum(1 for r in results if r.get('_flags'))} flagged)")
    return output_path
