#!/usr/bin/env python3
# create_weld_template.py
# Run from project_pablo/ directory to inspect the Weld template structure
 
import openpyxl
 
template_path = "files/Weld 1 Mail Survey Data Entry Spreadsheet Template_2.xlsx"
 
wb = openpyxl.load_workbook(template_path, read_only=True)
ws = wb.active
 
print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}")
print(f"Max col: {ws.max_column}")
print()
print("First 5 rows:")
for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
    print(f"  Row {i+1}: {row}")
 
wb.close()